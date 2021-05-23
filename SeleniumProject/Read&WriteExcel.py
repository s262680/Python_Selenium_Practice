import openpyxl

#store the excel file location to a variable
xlPath="C:\TempForTesting\Book.xlsx"

#load the workbook and store it into a variable
workbook=openpyxl.load_workbook(xlPath)

#get the current sheet, only works when there is only 1 sheet in the doc
#sheet=workbook.active

#get first sheet by searching for specified name
sheet1=workbook["Sheet1"]

#find the max number of rows and columns in the sheet and store them into variables
rows=sheet1.max_row
cols=sheet1.max_column

print(rows)
print(cols)

#loop through all the cells and print the values
for r in range(1, rows+1):
    for c in range(1, cols+1):
        print(sheet1.cell(row=r, column=c).value, end="  ")
    print("")


#get the second empty sheet
sheet2=workbook["Sheet2"]

for r in range(1, 8):
    for c in range(1, 5):
        if r==1:
            sheet2.cell(row=r, column=c).value = "header"
        else:
            sheet2.cell(row=r, column=c).value = str(r)+str(c)

workbook.save(xlPath)
