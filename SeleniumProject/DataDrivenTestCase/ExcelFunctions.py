import openpyxl

#A function that returns number of row in an excel sheet by entering file path and sheet name as parameter
def getRowCount(filePath, sheetName):
    workbook=openpyxl.load_workbook(filePath)
    sheet=workbook[sheetName]
    return(sheet.max_row)

#A function that returns number of column in an excel sheet by entering file path and sheet name as parameter
def getColumnCount(filePath,sheetName):
    workbook=openpyxl.load_workbook(filePath)
    sheet=workbook[sheetName]
    return(sheet.max_column)

#A function that returns data from a specific cell in an excel sheet
# by entering file path, sheet name row number and column number as parameter
def readData(filePath,sheetName, rowNum, columnNum):
    workbook=openpyxl.load_workbook(filePath)
    sheet=workbook[sheetName]
    return sheet.cell(row=rowNum, column=columnNum).value

#A function that write new data to a specific cell in an excel sheet
# by entering file path, sheet name row number, column number and new data as parameter
def writeData(filePath,sheetName, rowNum, columnNum, newData):
    workbook=openpyxl.load_workbook(filePath)
    sheet=workbook[sheetName]
    sheet.cell(row=rowNum, column=columnNum).value=newData
    workbook.save(filePath)

#A function that will print all data in a specific sheet by entering file path and sheet name as parameter
def printAllData(filePath,sheetName):
    workbook = openpyxl.load_workbook(filePath)
    sheet = workbook[sheetName]
    rows = sheet.max_row
    cols = sheet.max_column
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            print(sheet.cell(row=r, column=c).value, end="  ")
        print("")


#test runs
#path="C:\TempForTesting\Book.xlsx"

#print(getRowCount(path, "Sheet1"))

#print(readData(path, "Sheet1", 2, 3))


#writeData(path, "Sheet2", 4, 5, "newdata")
#print(readData(path, "Sheet2", 4, 5))

#printAllData(path, "Sheet1")

